from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl


def norm_text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_upper(v: Any) -> str:
    s = norm_text(v).upper()
    s = re.sub(r"[^A-Z0-9# ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def is_number(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    s = str(v).strip()
    if s == "":
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def to_num(v: Any) -> float | None:
    if not is_number(v):
        return None
    return float(str(v).strip())


def extract_spray_group(maint_text: str) -> str:
    m = re.search(r"\bSPRAY\s+DRAINS?\s+([A-Z0-9-]+)\b", maint_text)
    if m:
        return m.group(1)
    toks = maint_text.split(" ")
    ln = len(toks)
    for i in range(ln - 1):
        if toks[i] == "SPRAY" and toks[i + 1] in ("DRAIN", "DRAINS"):
            for j in range(i + 2, ln):
                t = toks[j]
                if t and t not in ("DRAIN", "DRAINS"):
                    return t
    return ""


def first_token(v: str) -> str:
    s = norm_upper(v)
    return s.split(" ")[0] if s else ""


def strip_segment_suffix(name: str) -> str:
    s = norm_text(name)
    s = re.sub(r"\s*\([^)]+\)\s*$", "", s)
    return s.strip()


def canonical_drain_asset_id(name: str) -> str:
    s = strip_segment_suffix(name)
    s = re.sub(r"^\s*drain\s+", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def find_header_row(ws, max_scan: int = 120) -> int | None:
    max_row = min(ws.max_row, max_scan)
    max_col = min(ws.max_column, 80)
    for r in range(1, max_row + 1):
        vals = [norm_upper(ws.cell(r, c).value) for c in range(1, max_col + 1)]
        has_mi = "MI" in vals or "MI #" in vals
        has_maint = "MAINTITEM TEXT" in vals or "LOCATION GROUPING" in vals
        if has_mi and has_maint:
            return r
    return None


def find_col(header_vals: list[str], names: list[str]) -> int | None:
    wanted = set(names)
    for i, h in enumerate(header_vals, start=1):
        if h in wanted:
            return i
    return None


def work_mapping(work_path: Path) -> dict[str, dict[str, str]]:
    wb = openpyxl.load_workbook(work_path, data_only=True)
    mapping: dict[str, dict[str, str]] = {}
    for ws in wb.worksheets:
        po = ""
        for c in range(1, min(20, ws.max_column) + 1):
            v = norm_text(ws.cell(1, c).value)
            if v:
                po = v
                break
        hr = find_header_row(ws)
        if hr is None:
            continue
        max_col = min(ws.max_column, 80)
        headers = [norm_upper(ws.cell(hr, c).value) for c in range(1, max_col + 1)]
        mi_col = find_col(headers, ["MI", "MI #"]) or 2
        maint_col = find_col(headers, ["MAINTITEM TEXT", "LOCATION GROUPING"])
        wo_col = find_col(headers, ["WORK ORDER", "WORK ORDERS", "WORK\nORDERS", "WORKORDER", "WORKORDERS"])
        if maint_col is None:
            continue
        for r in range(hr + 1, ws.max_row + 1):
            mi = norm_text(ws.cell(r, mi_col).value)
            maint = norm_upper(ws.cell(r, maint_col).value)
            wo_override = norm_text(ws.cell(r, wo_col).value) if wo_col else ""
            if not mi or not maint:
                continue
            grp = extract_spray_group(maint)
            if not grp:
                continue
            wo = wo_override if wo_override else mi
            mapping[grp] = {"work_order": wo, "purchase_order": po}
    return mapping


def parse_drain_job_workbooks(job_paths: list[Path], wo_map: dict[str, dict[str, str]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for fp in job_paths:
        wb = openpyxl.load_workbook(fp, data_only=True)
        for ws in wb.worksheets:
            catchment = ""
            for r in range(1, ws.max_row + 1):
                name = norm_text(ws.cell(r, 1).value)
                b = ws.cell(r, 2).value
                c = ws.cell(r, 3).value
                d = ws.cell(r, 4).value
                if not name and b is None and c is None and d is None:
                    continue
                if name and ("DRAIN" in norm_upper(name)) and ("START" in norm_upper(b) or "END" in norm_upper(c)):
                    continue
                if name and not is_number(b) and not is_number(c) and "CATCHMENT" in norm_upper(name):
                    catchment = name
                    continue
                start_m = to_num(b)
                end_m = to_num(c)
                total_m = to_num(d)
                if not name or start_m is None or end_m is None:
                    continue
                qty_km = (total_m if total_m is not None else max(0.0, end_m - start_m)) / 1000.0
                item = f"{name} ({int(start_m)}-{int(end_m)})"
                base = strip_segment_suffix(name)
                asset_id = canonical_drain_asset_id(name)
                grp = first_token(base)
                wo = ""
                po = ""
                if grp in wo_map:
                    wo = wo_map[grp]["work_order"]
                    po = wo_map[grp]["purchase_order"] if wo else ""
                job_key = f"drain|{ws.title}|{catchment}|{item}"
                if job_key in seen:
                    continue
                seen.add(job_key)
                out.append(
                    {
                        "job_key": job_key,
                        "module": "drain",
                        "asset_type": "drain",
                        "asset_id": asset_id,
                        "work_order": wo,
                        "purchase_order": po,
                        "status": "pending",
                        "scheduled_date": "",
                        "description": f"{item} / {ws.title} / {catchment}".strip(" /"),
                        "meta": {
                            "source": fp.name,
                            "sheet": ws.title,
                            "catchment": catchment,
                            "drain": name,
                            "start_m": start_m,
                            "end_m": end_m,
                            "qty_km": round(qty_km, 3),
                        },
                    }
                )
    return out


def parse_non_drain_work_rows(work_path: Path) -> list[dict[str, Any]]:
    modules = {
        "NOXIOUS WEEDS": ("weeds", "Noxious Weeds"),
        "FIRE ZONE": ("fire", "Fire Zone"),
        "MTN ACCESS TRACKS": ("tracks", "Mtn Access Tracks"),
    }
    wb = openpyxl.load_workbook(work_path, data_only=True)
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for ws in wb.worksheets:
        key = norm_upper(ws.title)
        if key not in modules:
            continue
        module_id, label = modules[key]
        po = ""
        for c in range(1, min(20, ws.max_column) + 1):
            v = norm_text(ws.cell(1, c).value)
            if v:
                po = v
                break
        max_col = min(ws.max_column, 80)
        r = 1
        while r <= ws.max_row:
            row_vals = [norm_upper(ws.cell(r, c).value) for c in range(1, max_col + 1)]
            is_header = ("MI" in row_vals or "MI #" in row_vals) and "MAINTITEM TEXT" in row_vals
            if not is_header:
                r += 1
                continue
            mi_col = find_col(row_vals, ["MI", "MI #"]) or 2
            suburb_col = find_col(row_vals, ["SUBURB TOWN", "SUBURB/TOWN"])
            call_col = find_col(row_vals, ["CALL DATE"])
            wo_col = find_col(row_vals, ["WORK ORDER", "WORK ORDERS", "WORK\nORDERS", "WORKORDER", "WORKORDERS"])
            rr = r + 1
            while rr <= ws.max_row:
                vals = [ws.cell(rr, c).value for c in range(1, max_col + 1)]
                hdr_check = [norm_upper(v) for v in vals]
                if ("MI" in hdr_check or "MI #" in hdr_check) and "MAINTITEM TEXT" in hdr_check:
                    break
                mi = norm_text(ws.cell(rr, mi_col).value)
                wo_override = norm_text(ws.cell(rr, wo_col).value) if wo_col else ""
                wo = wo_override if wo_override else mi
                if wo == "":
                    rr += 1
                    continue
                location = norm_text(ws.cell(rr, suburb_col).value) if suburb_col else ""
                call_date = norm_text(ws.cell(rr, call_col).value) if call_col else ""
                job_key = f"{module_id}|{ws.title}|{wo}"
                if job_key not in seen:
                    seen.add(job_key)
                    out.append(
                        {
                            "job_key": job_key,
                            "module": module_id,
                            "asset_type": "",
                            "asset_id": "",
                            "work_order": wo,
                            "purchase_order": po,
                            "status": "pending",
                            "scheduled_date": "",
                            "description": location if location else label,
                            "meta": {
                                "source": work_path.name,
                                "sheet": ws.title,
                                "location": location,
                                "call_date": call_date,
                            },
                        }
                    )
                rr += 1
            r = rr
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Build ILS V3 jobs payload from XLSX bundle.")
    ap.add_argument("--work", required=True, help="WO/PO workbook path")
    ap.add_argument("--jobs", nargs="+", required=True, help="One or more drain job workbooks")
    args = ap.parse_args()
    work = Path(args.work).resolve()
    jobs = [Path(p).resolve() for p in args.jobs]
    if not work.exists():
        print(json.dumps({"ok": False, "error": "work_file_missing"}))
        return 1
    for jp in jobs:
        if not jp.exists():
            print(json.dumps({"ok": False, "error": "job_file_missing", "file": str(jp)}))
            return 1
    try:
        wo_map = work_mapping(work)
        drain_jobs = parse_drain_job_workbooks(jobs, wo_map)
        other_jobs = parse_non_drain_work_rows(work)
        rows = drain_jobs + other_jobs
        print(
            json.dumps(
                {
                    "ok": True,
                    "rows": rows,
                    "counts": {
                        "drain_rows": len(drain_jobs),
                        "other_rows": len(other_jobs),
                        "total_rows": len(rows),
                        "work_map_groups": len(wo_map),
                    },
                },
                ensure_ascii=False,
            )
        )
        return 0
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
